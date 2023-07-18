import os

import cv2 as cv
import matplotlib.pyplot as plt
import numpy as np

from flask import Flask, render_template, request, redirect, send_file
#from werkzeug import secure_filename
from werkzeug.utils import secure_filename
from werkzeug.datastructures import  FileStorage
f = None
Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS = 15.18
Quantity_TOTAL_RHEL_LARGE_LICENSES = 89.28
Quantity_TOTAL_RHEL_SMALL_LICENSES = 119.04
Quantity_TOTAL_EGRESS_GB = 407.24
Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS = 1770.91
Quantity_MAX_VCPU = 4633.6
Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS = 6873.48
Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS = 6802.08
Quantity_TOTAL_WINDOWS_LICENSES_HOURS = 6745.97
Quantity_MAX_RAM_GB = 17826.6
Precio_Storage_025 = 0.000068
Precio_Storage_4 = 0.000219
Precio_Storage_2 = 0.000164
Precio_Storage_10 = 0.000658

#def wholeFunction():
#  print(Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS)
#  print(type(Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS))

def wholeFunction():
  #Librerias
  from openpyxl import Workbook
  from openpyxl.styles import Font
  from openpyxl.styles import PatternFill
  import openpyxl
  from datetime import datetime
  from openpyxl.writer.excel import save_virtual_workbook

  #import numpy as np

  #Constantes
  wb = openpyxl.load_workbook(f.filename)
  sheet = wb.active
  number_of_rows = 0
  DayOn = ''
  MonthArray = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Setiembre', 'Octubre', 'Noviembre', 'Diciembre']
  #Ingreso de constantes para las formulas
  Total_Ram_MB = 0
  Total_Ram_GB = 0 #Me parece que este dato no se utiliza
  Total_Vcpu = 0
  Total_Storage_MB = 0
  Total_Storage_GB = 0
  Total_Costo_Storage = 0
  Total_Costo_Storage_Unkown = 0
  Total_Costo_Egress_Unit = 0
  Total_Precio_Final_Mas_MAX_RAM_GB = 0
  Total_Precio_Final_Mas_MAX_VCPU = 0
  Total_Costo_VCPU = 0
  Total_Calculos_Storage_025 = 0
  Total_Calculos_Storage_4 = 0
  Total_Calculos_Storage_2 = 0
  Total_Calculos_Storage_10 = 0
  Total_Porcentaje_Storage_025 = 0
  Total_Porcentaje_Storage_4 = 0
  Total_Porcentaje_Storage_2 = 0
  Total_Porcentaje_Storage_10 = 0
  Total_Precio_Final_Storage_025 = 0
  Total_Precio_Final_Storage_4 = 0
  Total_Precio_Final_Storage_2 = 0
  Total_Precio_Final_Storage_10 = 0
  Total_Calculos_Licencias_Windows = 0
  Total_Calculos_Licencias_RHS = 0
  Total_Calculos_Licencias_RHL = 0
  Total_Porcentaje_Licencias_Windows = 0
  Total_Porcentaje_Licencias_RHS = 0
  Total_Porcentaje_Licencias_RHL = 0
  Total_Precio_Final_Licencias_Windows = 0
  Total_Precio_Final_Licencias_RHS = 0
  Total_Precio_Final_Licencias_RHL = 0
  Total_Precio_Final_Por_Máquina = 0
  Suma_Precio_Final = 0
  #datos input

  #Sacar valores de las columnas y filas / no se todavia como funciona
  def CellVal(r, c):
      return sheet.cell(row=r, column=c).value

  #Sacar el numero de filas
  for row in range(1,sheet.max_row):
      if(sheet.cell(row,1).value is not None):
              number_of_rows = number_of_rows + 1
  number_of_rows = number_of_rows +2 #Todavia no entiendo el porque se debe sumar 2

  #Crear la nueva hoja
  sheet2 = wb.create_sheet('output')

  #Poner nombres a la primera fila
  #Datos input
  sheet2['A1'] = 'ID de VM'
  sheet2['A1'].font = Font(color='000000', bold=True)
  #----datos nuevos 19 junio--------
  sheet2['B1'] = 'Container Name'
  sheet2['B1'].font = Font(color='000000', bold=True)
  sheet2['C1'] = 'IP Address'
  sheet2['C1'].font = Font(color='000000', bold=True)
  #----datos nuevos 19 junio--------
  sheet2['D1'] = 'Licencia'
  sheet2['D1'].font = Font(color='000000', bold=True)
  sheet2['E1'] = 'Dias prendido'
  sheet2['E1'].font = Font(color='000000', bold=True)
  sheet2['F1'] = 'Status de la VM'
  sheet2['F1'].font = Font(color='000000', bold=True)
  sheet2['G1'] = 'Megabytes de RAM'
  sheet2['G1'].font = Font(color='000000', bold=True)
  sheet2['H1'] = 'Gigabytes de RAM'
  sheet2['H1'].font = Font(color='000000', bold=True)
  sheet2['I1'] = 'Cantidad de VCPUs'
  sheet2['I1'].font = Font(color='000000', bold=True)
  sheet2['J1'] = 'Megabytes de Storage'
  sheet2['J1'].font = Font(color='000000', bold=True)
  sheet2['K1'] = 'Gigabytes de Storage'
  sheet2['K1'].font = Font(color='000000', bold=True)
  sheet2['L1'] = 'Storage Profile'
  sheet2['L1'].font = Font(color='000000', bold=True)
  #DATOS OUTPUT
  #RAM
  sheet2['M1'] = 'Costo GB RAM Unit  $0.0124/GB-hr'
  sheet2['M1'].font = Font(color='ffffff', bold=True)
  sheet2['M1'].fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  sheet2['N1'] = 'Porcentaje del total de RAM en GB'
  sheet2['N1'].font = Font(color='ffffff', bold=True)
  sheet2['N1'].fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  sheet2['O1'] = 'Diferenciador'
  sheet2['O1'].font = Font(color='ffffff', bold=True)
  sheet2['O1'].fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  sheet2['P1'] = 'Precio Final - MAX_RAM_GB'
  sheet2['P1'].font = Font(color='ffffff', bold=True)
  sheet2['P1'].fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  #VCPU
  sheet2['Q1'] = 'COSTO VCPU 0.0111/hr'
  sheet2['Q1'].font = Font(color='ffffff', bold=True)
  sheet2['Q1'].fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  sheet2['R1'] = 'Porcentaje del total de VCPU en GB'
  sheet2['R1'].font = Font(color='ffffff', bold=True)
  sheet2['R1'].fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  sheet2['S1'] = 'Diferenciador'
  sheet2['S1'].font = Font(color='ffffff', bold=True)
  sheet2['S1'].fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  sheet2['T1'] = 'Precio Final - MAX_VCPU'
  sheet2['T1'].font = Font(color='ffffff', bold=True)
  sheet2['T1'].fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  #Storage Egress
  sheet2['U1'] = 'Porcentaje del total de Storage'
  sheet2['U1'].font = Font(color='ffffff', bold=True)
  sheet2['U1'].fill = PatternFill(start_color='2b72d6', end_color='2b72d6',fill_type = "solid")
  sheet2['V1'] = 'COSTO EGRESS Unit 0.09/GB'
  sheet2['V1'].font = Font(color='ffffff', bold=True)
  sheet2['V1'].fill = PatternFill(start_color='2b72d6', end_color='2b72d6',fill_type = "solid")
  #Storage 0.25 IOPS
  sheet2['W1'] = 'Calculos de Storage 0.25 IOPS'
  sheet2['W1'].font = Font(color='ffffff', bold=True)
  sheet2['W1'].fill = PatternFill(start_color='71b2e3', end_color='71b2e3',fill_type = "solid")
  sheet2['X1'] = 'Porcentaje de Storage 0.25 IOPS'
  sheet2['X1'].font = Font(color='ffffff', bold=True)
  sheet2['X1'].fill = PatternFill(start_color='71b2e3', end_color='71b2e3',fill_type = "solid")
  sheet2['Y1'] = 'Precio final de Storage 0.25 IOPS'
  sheet2['Y1'].font = Font(color='ffffff', bold=True)
  sheet2['Y1'].fill = PatternFill(start_color='71b2e3', end_color='71b2e3',fill_type = "solid")
  #Storage 4 IOPS
  sheet2['Z1'] = 'Calculos de Storage 4 IOPS'
  sheet2['Z1'].font = Font(color='ffffff', bold=True)
  sheet2['Z1'].fill = PatternFill(start_color='707de0', end_color='707de0',fill_type = "solid")
  sheet2['AA1'] = 'Porcentaje de Storage 4 IOPS'
  sheet2['AA1'].font = Font(color='ffffff', bold=True)
  sheet2['AA1'].fill = PatternFill(start_color='707de0', end_color='707de0',fill_type = "solid")
  sheet2['AB1'] = 'Precio final de Storage 4 IOPS'
  sheet2['AB1'].font = Font(color='ffffff', bold=True)
  sheet2['AB1'].fill = PatternFill(start_color='707de0', end_color='707de0',fill_type = "solid")
  #Storage 2 IOPS
  sheet2['AC1'] = 'Calculos de Storage 2 IOPS'
  sheet2['AC1'].font = Font(color='ffffff', bold=True)
  sheet2['AC1'].fill = PatternFill(start_color='8770e0', end_color='8770e0',fill_type = "solid")
  sheet2['AD1'] = 'Porcentaje de Storage 2 IOPS'
  sheet2['AD1'].font = Font(color='ffffff', bold=True)
  sheet2['AD1'].fill = PatternFill(start_color='8770e0', end_color='8770e0',fill_type = "solid")
  sheet2['AE1'] = 'Precio final de Storage 2 IOPS'
  sheet2['AE1'].font = Font(color='ffffff', bold=True)
  sheet2['AE1'].fill = PatternFill(start_color='8770e0', end_color='8770e0',fill_type = "solid")
  #Storage 10 IOPS
  sheet2['AF1'] = 'Calculos de Storage 10 IOPS'
  sheet2['AF1'].font = Font(color='ffffff', bold=True)
  sheet2['AF1'].fill = PatternFill(start_color='e07097', end_color='e07097',fill_type = "solid")
  sheet2['AG1'] = 'Porcentaje de Storage 10 IOPS'
  sheet2['AG1'].font = Font(color='ffffff', bold=True)
  sheet2['AG1'].fill = PatternFill(start_color='e07097', end_color='e07097',fill_type = "solid")
  sheet2['AH1'] = 'Precio final de Storage 10 IOPS'
  sheet2['AH1'].font = Font(color='ffffff', bold=True)
  sheet2['AH1'].fill = PatternFill(start_color='e07097', end_color='e07097',fill_type = "solid")
  #Licencias Windows
  sheet2['AI1'] = 'Calculos de Licencias Windows'
  sheet2['AI1'].font = Font(color='ffffff', bold=True)
  sheet2['AI1'].fill = PatternFill(start_color='bfd180', end_color='bfd180',fill_type = "solid")
  sheet2['AJ1'] = 'Porcentaje de Licencias Windows'
  sheet2['AJ1'].font = Font(color='ffffff', bold=True)
  sheet2['AJ1'].fill = PatternFill(start_color='bfd180', end_color='bfd180',fill_type = "solid")
  sheet2['AK1'] = 'Precio final de Licencias Windows'
  sheet2['AK1'].font = Font(color='ffffff', bold=True)
  sheet2['AK1'].fill = PatternFill(start_color='bfd180', end_color='bfd180',fill_type = "solid")
  #Licencias RHS
  sheet2['AL1'] = 'Calculos de Licencias Red Hat Small'
  sheet2['AL1'].font = Font(color='ffffff', bold=True)
  sheet2['AL1'].fill = PatternFill(start_color='9cd180', end_color='9cd180',fill_type = "solid")
  sheet2['AM1'] = 'Porcentaje de Licencias Red Hat Small'
  sheet2['AM1'].font = Font(color='ffffff', bold=True)
  sheet2['AM1'].fill = PatternFill(start_color='9cd180', end_color='9cd180',fill_type = "solid")
  sheet2['AN1'] = 'Precio final de Licencias Red Hat Small'
  sheet2['AN1'].font = Font(color='ffffff', bold=True)
  sheet2['AN1'].fill = PatternFill(start_color='9cd180', end_color='9cd180',fill_type = "solid")
  #Licencias RHL
  sheet2['AO1'] = 'Calculos de Licencias Red Hat Large'
  sheet2['AO1'].font = Font(color='ffffff', bold=True)
  sheet2['AO1'].fill = PatternFill(start_color='d18088', end_color='d18088',fill_type = "solid")
  sheet2['AP1'] = 'Porcentaje de Licencias Red Hat Large'
  sheet2['AP1'].font = Font(color='ffffff', bold=True)
  sheet2['AP1'].fill = PatternFill(start_color='d18088', end_color='d18088',fill_type = "solid")
  sheet2['AQ1'] = 'Precio final de Licencias Red Hat Large'
  sheet2['AQ1'].font = Font(color='ffffff', bold=True)
  sheet2['AQ1'].fill = PatternFill(start_color='d18088', end_color='d18088',fill_type = "solid")
  sheet2['AR1'] = 'PRECIO FINAL'
  sheet2['AR1'].font = Font(color='ffffff', bold=True)
  sheet2['AR1'].fill = PatternFill(start_color='7bb585', end_color='7bb585',fill_type = "solid")
  #Agregando la plabra Total
  sheet2.cell(row=(number_of_rows), column=6).value = 'TOTAL'
  sheet2.cell(row=(number_of_rows), column=6).font = Font(color='FF0000', bold=True)

  #------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
  #Ingreso de datos del Portal de IBM Cloud
  New_Row = number_of_rows + 3 #Nueva fila de referencia
  #Títulos
  #sheet2.cell(row=New_Row, column=1).value = 'Quantity'
  #sheet2.cell(row=New_Row, column=1).font = Font(color='006aff', bold=True)
  sheet2.cell(row=New_Row, column=2).value = 'Unit'
  sheet2.cell(row=New_Row, column=2).font = Font(color='006aff', bold=True)
  sheet2.cell(row=New_Row, column=3).value = 'Unit ID'
  sheet2.cell(row=New_Row, column=3).font = Font(color='006aff', bold=True)
  sheet2.cell(row=New_Row, column=4).value = 'Cost (USD)'
  sheet2.cell(row=New_Row, column=4).font = Font(color='006aff', bold=True)
  #Quantity --- Se saca del portal de IBM Cloud

  #Unit
  sheet2.cell(row=New_Row+1, column=2).value = 'Gigabyte-Hour'
  sheet2.cell(row=New_Row+2, column=2).value = 'Instance-Hour'
  sheet2.cell(row=New_Row+3, column=2).value = 'Instance-Hour'
  sheet2.cell(row=New_Row+4, column=2).value = 'Gigabyte Transmitted Outbound'
  sheet2.cell(row=New_Row+5, column=2).value = 'Gigabyte-Hour'
  sheet2.cell(row=New_Row+6, column=2).value = 'Instance'
  sheet2.cell(row=New_Row+7, column=2).value = 'Gigabyte-Hour'
  sheet2.cell(row=New_Row+8, column=2).value = 'Gigabyte-Hour'
  sheet2.cell(row=New_Row+9, column=2).value = 'Instance-Hours'
  sheet2.cell(row=New_Row+10, column=2).value = 'Instance'
  #Unit ID
  sheet2.cell(row=New_Row+1, column=3).value = 'TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS'
  sheet2.cell(row=New_Row+2, column=3).value = 'TOTAL_RHEL_LARGE_LICENSES'
  sheet2.cell(row=New_Row+3, column=3).value = 'TOTAL_RHEL_SMALL_LICENSES'
  sheet2.cell(row=New_Row+4, column=3).value = 'TOTAL_EGRESS_GB'
  sheet2.cell(row=New_Row+5, column=3).value = 'TOTAL_STORAGE_TWO_IOPS_GB_HOURS'
  sheet2.cell(row=New_Row+6, column=3).value = 'MAX_VCPU'
  sheet2.cell(row=New_Row+7, column=3).value = 'TOTAL_STORAGE_TEN_IOPS_GB_HOURS'
  sheet2.cell(row=New_Row+8, column=3).value = 'TOTAL_STORAGE_FOUR_IOPS_GB_HOURS'
  sheet2.cell(row=New_Row+9, column=3).value = 'TOTAL_WINDOWS_LICENSES_HOURS'
  sheet2.cell(row=New_Row+10, column=3).value = 'MAX_RAM_GB'
  #Los datos que se necesita para los datos que no conozco de VCPU y RAM
  #Cost --- Se saca del portal de IBM Cloud
  sheet2.cell(row=New_Row+1, column=4).value = Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS
  sheet2.cell(row=New_Row+2, column=4).value = Quantity_TOTAL_RHEL_LARGE_LICENSES
  sheet2.cell(row=New_Row+3, column=4).value = Quantity_TOTAL_RHEL_SMALL_LICENSES
  sheet2.cell(row=New_Row+4, column=4).value = Quantity_TOTAL_EGRESS_GB
  sheet2.cell(row=New_Row+5, column=4).value = Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS
  sheet2.cell(row=New_Row+6, column=4).value = Quantity_MAX_VCPU

  #sheet2.cell(row=New_Row+6, column=5).value = sheet2.cell(row=New_Row+6, column=4).value - 

  sheet2.cell(row=New_Row+7, column=4).value = Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS
  sheet2.cell(row=New_Row+8, column=4).value = Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS
  sheet2.cell(row=New_Row+9, column=4).value = Quantity_TOTAL_WINDOWS_LICENSES_HOURS
  sheet2.cell(row=New_Row+10, column=4).value = Quantity_MAX_RAM_GB

  #------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

  #INGRESO DE CADA CELDA POR FILA CORRESPONDIENTE
  #DATOS INPUT
  #Ingreso del ID
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=1).value = CellVal(m, 2)
  
  
  #----datos nuevos 19 junio--------
  #Ingreso del container name
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=2).value = CellVal(m, 8)
  #Ingreso del ip address
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=3).value = CellVal(m, 12)
  #Ingreso de la licencia
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=4).value = CellVal(m, 10)
  #Ingreso del status de la VM
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=6).value = CellVal(m, 29)
  #Ingreso de la # dias prendido
  for m in range(2, number_of_rows):
      #var = CellVal(m, 40)
      #var = var.replace('-', '/')
      #for letter in var:
      #  if letter == 'T':
      #    break
      #  else:
      #    DayOn = DayOn + letter
      Today = datetime.today().strftime('%Y/%m/%d')
      Today = datetime.strptime(Today, '%Y/%m/%d')
      #DayOn = datetime.strptime(DayOn, "%Y/%m/%d")
      #Today = datetime.strptime(Today, "%Y/%m/%d")
      #print(Today)
      #print(type(Today))
      #print('---------------')
      #print(DayOn)
      #print(type(DayOn))
      #print('---------------')
      #subs = Today - DayOn
      
      #sheet2.cell(row=m, column=3).value = subs
      #DayOn = ''
      if (sheet2.cell(row=m, column=6).value) == 'POWERED_OFF':
        sheet2.cell(row=m, column=5).value = 0
      else:
        if(Today.month - 1 == 2):
          sheet2.cell(row=m, column=5).value = 28
        else:
          sheet2.cell(row=m, column=5).value = 31
      #///OJO//// aca tengo que saber la ultima vez que se prendio la maquina
  #sheet2.cell(row=36, column=3).value = 11
  #sheet2.cell(row=98, column=3).value = 1
  # Falta poner el número de días que esta prendido una máquina que fue creada luego del primero del mes que se quiere sacar el reporte
  #Ingreso de la RAM MB
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=7).value = CellVal(m, 22)
      Total_Ram_MB = Total_Ram_MB + CellVal(m, 22)
  sheet2.cell(row=(number_of_rows), column=7).value = Total_Ram_MB
  sheet2.cell(row=(number_of_rows), column=7).font = Font(color='000000', bold=True)
  #Ingreso de la # VCPU's
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=9).value = CellVal(m, 26)
      Total_Vcpu = Total_Vcpu + CellVal(m, 26)
  sheet2.cell(row=(number_of_rows), column=9).value = Total_Vcpu
  sheet2.cell(row=(number_of_rows), column=9).font = Font(color='000000', bold=True)
  #Ingreso de la Storage MB
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=10).value = CellVal(m, 48)
      Total_Storage_MB = Total_Storage_MB + CellVal(m, 48)
  sheet2.cell(row=(number_of_rows), column=10).value = Total_Storage_MB
  sheet2.cell(row=(number_of_rows), column=10).font = Font(color='000000', bold=True)
  #Ingreso del storage profile
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=12).value = CellVal(m, 30)
  #DATOS OUTPUT
  #Sacar valores de las columnas y filas de la oja output/ no se todavia como funciona
  

  def CellVal(r, c):
      return sheet2.cell(row=r, column=c).value
  #Ingreso de la RAM GB
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=8).value = ((sheet2.cell(row=m, column=7).value)/1024)
      Total_Ram_GB = Total_Ram_GB + ((sheet2.cell(row=m, column=7).value)/1024)
  sheet2.cell(row=(number_of_rows), column=8).value = Total_Ram_GB
  sheet2.cell(row=(number_of_rows), column=8).font = Font(color='000000', bold=True)
  #Ingreso del Storage GB
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=11).value = ((sheet2.cell(row=m, column=10).value)/1024)
      Total_Storage_GB = Total_Storage_GB + ((sheet2.cell(row=m, column=10).value)/1024)
  sheet2.cell(row=(number_of_rows), column=11).value = Total_Storage_GB
  sheet2.cell(row=(number_of_rows), column=11).font = Font(color='000000', bold=True)
  #Ingreso de Costo RAM unitario
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=13).value = ((sheet2.cell(row=m, column=8).value)*0.0124*24*(sheet2.cell(row=m, column=5).value))
      Total_Costo_Storage = Total_Costo_Storage + ((sheet2.cell(row=m, column=8).value)*0.0124*24*(sheet2.cell(row=m, column=5).value))
  sheet2.cell(row=(number_of_rows), column=13).value = Total_Costo_Storage
  sheet2.cell(row=New_Row+10, column=5).value = float(sheet2.cell(row=New_Row+10, column=4).value) - Total_Costo_Storage
  sheet2.cell(row=(number_of_rows), column=13).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=(number_of_rows), column=13).fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  #Ingreso del porcentaje RAM
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=14).value = sheet2.cell(row=m, column=13).value/sheet2.cell(row=(number_of_rows), column=13).value
      #varToStr = str(sheet2.cell(row=m, column=12).value)
      #varToStr = varToStr + '%'
      #sheet2.cell(row=m, column=12).value = varToStr
      #varToStr = ''
  #Ingreso del dato que no conozco de RAM
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=15).value = sheet2.cell(row=m, column=13).value/sheet2.cell(row=(number_of_rows), column=13).value * sheet2.cell(row=New_Row+10, column=5).value
      Total_Costo_Storage_Unkown = Total_Costo_Storage_Unkown + sheet2.cell(row=m, column=15).value
  sheet2.cell(row=(number_of_rows), column=15).value = Total_Costo_Storage_Unkown
  sheet2.cell(row=(number_of_rows), column=15).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=(number_of_rows), column=15).fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  #Ingreso de Precio Final - MAX_RAM_GB
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=16).value = sheet2.cell(row=m, column=13).value + sheet2.cell(row=m, column=15).value
      Total_Precio_Final_Mas_MAX_RAM_GB = Total_Precio_Final_Mas_MAX_RAM_GB + sheet2.cell(row=m, column=16).value
  sheet2.cell(row=(number_of_rows), column=16).value = Total_Precio_Final_Mas_MAX_RAM_GB
  sheet2.cell(row=(number_of_rows), column=16).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=(number_of_rows), column=16).fill = PatternFill(start_color='eb660e', end_color='eb660e',fill_type = "solid")
  #Ingreso de COSTO VCPU 
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=17).value = sheet2.cell(row=m, column=9).value * 0.0111 * 24 * sheet2.cell(row=m, column=5).value
      Total_Costo_VCPU = Total_Costo_VCPU + sheet2.cell(row=m, column=17).value
  sheet2.cell(row=number_of_rows, column=17).value = Total_Costo_VCPU
  sheet2.cell(row=New_Row+6, column=5).value = float(sheet2.cell(row=New_Row+6, column=4).value) - Total_Costo_VCPU
  sheet2.cell(row=number_of_rows, column=17).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=17).fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  #Ingreso del porcentaje VCPU
  for m in range(2, number_of_rows): #posiblemente falle aquí
      sheet2.cell(row=m, column=18).value = sheet2.cell(row=m, column=17).value/sheet2.cell(row=(number_of_rows), column=17).value
  #Ingreso del dato que no conozco de VCPU
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=19).value = sheet2.cell(row=m, column=17).value/sheet2.cell(row=(number_of_rows), column=17).value * sheet2.cell(row=New_Row+6, column=5).value
      Total_Costo_Storage_Unkown = Total_Costo_Storage_Unkown + sheet2.cell(row=m, column=19).value
  sheet2.cell(row=(number_of_rows), column=19).value = Total_Costo_Storage_Unkown
  sheet2.cell(row=(number_of_rows), column=19).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=19).fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  #Ingreso de Precio Final - MAX_VCPU
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=20).value = sheet2.cell(row=m, column=17).value + sheet2.cell(row=m, column=19).value
      Total_Precio_Final_Mas_MAX_VCPU = Total_Precio_Final_Mas_MAX_VCPU + sheet2.cell(row=m, column=20).value
  sheet2.cell(row=(number_of_rows), column=20).value = Total_Precio_Final_Mas_MAX_VCPU
  sheet2.cell(row=(number_of_rows), column=20).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=20).fill = PatternFill(start_color='107d16', end_color='107d16',fill_type = "solid")
  #Ingreso del porcentaje del storage 
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=21).value = sheet2.cell(row=m, column=10).value / sheet2.cell(row=(number_of_rows), column=10).value
  #Ingreso del COSTO EGRESS Unit 0.09/GB
  for m in range(2, number_of_rows):
      sheet2.cell(row=m, column=22).value = sheet2.cell(row=m, column=10).value / sheet2.cell(row=(number_of_rows), column=10).value
      sheet2.cell(row=m, column=22).value = float(sheet2.cell(row=m, column=22).value) * float(Quantity_TOTAL_EGRESS_GB)
      Total_Costo_Egress_Unit = Total_Costo_Egress_Unit + sheet2.cell(row=m, column=22).value
  sheet2.cell(row=(number_of_rows), column=22).value = Total_Costo_Egress_Unit
  sheet2.cell(row=(number_of_rows), column=22).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=(number_of_rows), column=22).fill = PatternFill(start_color='2b72d6', end_color='2b72d6',fill_type = "solid")
  #Ingreso del calculo del costo de storage por perfil -------------
  for m in range(2, number_of_rows):
      if(sheet2.cell(row=m, column=12).value == '0.25 IOPS/GB'): #0.25
        sheet2.cell(row=m, column=23).value = sheet2.cell(row=m, column=8).value * Precio_Storage_025 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Storage_025 = Total_Calculos_Storage_025 + sheet2.cell(row=m, column=23).value
      elif (sheet2.cell(row=m, column=12).value == '4 IOPS/GB'): #4
        sheet2.cell(row=m, column=26).value = sheet2.cell(row=m, column=8).value * Precio_Storage_4 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Storage_4 = Total_Calculos_Storage_4 + sheet2.cell(row=m, column=26).value
      elif (sheet2.cell(row=m, column=12).value == '2 IOPS/GB'): #2
        sheet2.cell(row=m, column=29).value = sheet2.cell(row=m, column=8).value * Precio_Storage_2 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Storage_2 = Total_Calculos_Storage_2 + sheet2.cell(row=m, column=29).value
      elif (sheet2.cell(row=m, column=12).value == '10 IOPS/GB'): #10
        sheet2.cell(row=m, column=32).value = sheet2.cell(row=m, column=8).value * Precio_Storage_10 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Storage_10 = Total_Calculos_Storage_10 + sheet2.cell(row=m, column=32).value
  sheet2.cell(row=number_of_rows, column=23).value = Total_Calculos_Storage_025
  sheet2.cell(row=number_of_rows, column=23).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=23).fill = PatternFill(start_color='71b2e3', end_color='71b2e3',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=26).value = Total_Calculos_Storage_4
  sheet2.cell(row=number_of_rows, column=26).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=26).fill = PatternFill(start_color='707de0', end_color='707de0',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=29).value = Total_Calculos_Storage_2
  sheet2.cell(row=number_of_rows, column=29).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=29).fill = PatternFill(start_color='8770e0', end_color='8770e0',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=32).value = Total_Calculos_Storage_10
  sheet2.cell(row=number_of_rows, column=32).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=32).fill = PatternFill(start_color='e07097', end_color='e07097',fill_type = "solid")
  #Ingreso del porcentaje de storage por perfil
  for m in range(2, number_of_rows):
        if(sheet2.cell(row=m, column=12).value == '0.25 IOPS/GB'): #0.25
          sheet2.cell(row=m, column=24).value = sheet2.cell(row=m, column=23).value / Total_Calculos_Storage_025
          Total_Porcentaje_Storage_025 = Total_Porcentaje_Storage_025 + sheet2.cell(row=m, column=24).value
        elif (sheet2.cell(row=m, column=12).value == '4 IOPS/GB'): #4
          sheet2.cell(row=m, column=27).value = sheet2.cell(row=m, column=26).value / Total_Calculos_Storage_4
          Total_Porcentaje_Storage_4 = Total_Porcentaje_Storage_4 + sheet2.cell(row=m, column=27).value
        elif (sheet2.cell(row=m, column=12).value == '2 IOPS/GB'): #2
          sheet2.cell(row=m, column=30).value = sheet2.cell(row=m, column=29).value / Total_Calculos_Storage_2
          Total_Porcentaje_Storage_2 = Total_Porcentaje_Storage_2 + sheet2.cell(row=m, column=30).value
        elif (sheet2.cell(row=m, column=12).value == '10 IOPS/GB'): #10
          sheet2.cell(row=m, column=33).value = sheet2.cell(row=m, column=32).value / Total_Calculos_Storage_10
          Total_Porcentaje_Storage_10 = Total_Porcentaje_Storage_10 + sheet2.cell(row=m, column=33).value
  sheet2.cell(row=number_of_rows, column=24).value = Total_Porcentaje_Storage_025
  sheet2.cell(row=number_of_rows, column=24).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=24).fill = PatternFill(start_color='71b2e3', end_color='71b2e3',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=27).value = Total_Porcentaje_Storage_4
  sheet2.cell(row=number_of_rows, column=27).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=27).fill = PatternFill(start_color='707de0', end_color='707de0',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=30).value = Total_Porcentaje_Storage_2
  sheet2.cell(row=number_of_rows, column=30).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=30).fill = PatternFill(start_color='8770e0', end_color='8770e0',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=33).value = Total_Porcentaje_Storage_10
  sheet2.cell(row=number_of_rows, column=33).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=33).fill = PatternFill(start_color='e07097', end_color='e07097',fill_type = "solid")
  #Ingreso del precio final del storage por perfil
  for m in range(2, number_of_rows):
        if(sheet2.cell(row=m, column=12).value == '0.25 IOPS/GB'): #0.25
          sheet2.cell(row=m, column=25).value = sheet2.cell(row=m, column=23).value / Total_Calculos_Storage_025 
          sheet2.cell(row=m, column=25).value = float(sheet2.cell(row=m, column=25).value) * float(Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS)
          Total_Precio_Final_Storage_025 = Total_Precio_Final_Storage_025 + sheet2.cell(row=m, column=25).value
        elif (sheet2.cell(row=m, column=12).value == '4 IOPS/GB'): #4
          sheet2.cell(row=m, column=28).value = sheet2.cell(row=m, column=26).value / Total_Calculos_Storage_4 
          sheet2.cell(row=m, column=28).value = float(sheet2.cell(row=m, column=28).value) * float(Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS)
          Total_Precio_Final_Storage_4 = Total_Precio_Final_Storage_4 + sheet2.cell(row=m, column=28).value
        elif (sheet2.cell(row=m, column=12).value == '2 IOPS/GB'): #2
          sheet2.cell(row=m, column=31).value = sheet2.cell(row=m, column=29).value / Total_Calculos_Storage_2 
          sheet2.cell(row=m, column=31).value = float(sheet2.cell(row=m, column=31).value) * float(Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS)
          Total_Precio_Final_Storage_2 = Total_Precio_Final_Storage_2 + sheet2.cell(row=m, column=31).value
        elif (sheet2.cell(row=m, column=12).value == '10 IOPS/GB'): #10
          sheet2.cell(row=m, column=34).value = sheet2.cell(row=m, column=32).value / Total_Calculos_Storage_10 
          sheet2.cell(row=m, column=34).value = float(sheet2.cell(row=m, column=34).value) * float(Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS)
          Total_Precio_Final_Storage_10 = Total_Precio_Final_Storage_10 + sheet2.cell(row=m, column=34).value
  sheet2.cell(row=number_of_rows, column=25).value = Total_Precio_Final_Storage_025
  sheet2.cell(row=number_of_rows, column=25).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=25).fill = PatternFill(start_color='71b2e3', end_color='71b2e3',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=28).value = Total_Precio_Final_Storage_4
  sheet2.cell(row=number_of_rows, column=28).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=28).fill = PatternFill(start_color='707de0', end_color='707de0',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=31).value = Total_Precio_Final_Storage_2
  sheet2.cell(row=number_of_rows, column=31).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=31).fill = PatternFill(start_color='8770e0', end_color='8770e0',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=34).value = Total_Precio_Final_Storage_10
  sheet2.cell(row=number_of_rows, column=34).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=34).fill = PatternFill(start_color='e07097', end_color='e07097',fill_type = "solid")
  #Ingreso del calculo del costo por licencia
  for m in range(2, number_of_rows):
      if('Microsoft Windows ' in sheet2.cell(row=m, column=4).value): #windows
        sheet2.cell(row=m, column=35).value = sheet2.cell(row=m, column=9).value * 0.0193 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Licencias_Windows = Total_Calculos_Licencias_Windows + sheet2.cell(row=m, column=35).value
      elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value <= 4): #Red Hat
        sheet2.cell(row=m, column=38).value = sheet2.cell(row=m, column=9).value * 0.0193 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Licencias_RHS = Total_Calculos_Licencias_RHS + sheet2.cell(row=m, column=38).value
      elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value > 4): #windows
        sheet2.cell(row=m, column=41).value = sheet2.cell(row=m, column=9).value * 0.0193 * 24 * sheet2.cell(row=m, column=5).value
        Total_Calculos_Licencias_RHL = Total_Calculos_Licencias_RHL + sheet2.cell(row=m, column=41).value
  sheet2.cell(row=number_of_rows, column=35).value = Total_Calculos_Licencias_Windows
  sheet2.cell(row=number_of_rows, column=35).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=35).fill = PatternFill(start_color='bfd180', end_color='bfd180',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=38).value = Total_Calculos_Licencias_RHS
  sheet2.cell(row=number_of_rows, column=38).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=38).fill = PatternFill(start_color='9cd180', end_color='9cd180',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=41).value = Total_Calculos_Licencias_RHL
  sheet2.cell(row=number_of_rows, column=41).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=41).fill = PatternFill(start_color='d18088', end_color='d18088',fill_type = "solid")
  #Ingreso del porcentaje por licencia
  for m in range(2, number_of_rows):
      if('Microsoft Windows ' in sheet2.cell(row=m, column=4).value): #windows
        sheet2.cell(row=m, column=36).value = sheet2.cell(row=m, column=35).value / Total_Calculos_Licencias_Windows
        Total_Porcentaje_Licencias_Windows = Total_Porcentaje_Licencias_Windows + sheet2.cell(row=m, column=36).value
      elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value <= 4): #Red Hat
        sheet2.cell(row=m, column=39).value = sheet2.cell(row=m, column=38).value / Total_Calculos_Licencias_RHS
        Total_Porcentaje_Licencias_RHS = Total_Porcentaje_Licencias_RHS + sheet2.cell(row=m, column=39).value
      elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value > 4): #windows
        sheet2.cell(row=m, column=42).value = sheet2.cell(row=m, column=41).value / Total_Calculos_Licencias_RHL
        Total_Porcentaje_Licencias_RHL = Total_Porcentaje_Licencias_RHL + sheet2.cell(row=m, column=42).value
  sheet2.cell(row=number_of_rows, column=36).value = Total_Porcentaje_Licencias_Windows
  sheet2.cell(row=number_of_rows, column=36).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=36).fill = PatternFill(start_color='bfd180', end_color='bfd180',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=39).value = Total_Porcentaje_Licencias_RHS
  sheet2.cell(row=number_of_rows, column=39).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=39).fill = PatternFill(start_color='9cd180', end_color='9cd180',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=42).value = Total_Porcentaje_Licencias_RHL
  sheet2.cell(row=number_of_rows, column=42).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=42).fill = PatternFill(start_color='d18088', end_color='d18088',fill_type = "solid")
  #Ingreso del precio final por licencia
  for m in range(2, number_of_rows):
      if('Microsoft Windows ' in sheet2.cell(row=m, column=4).value): #windows
        sheet2.cell(row=m, column=37).value = float(sheet2.cell(row=m, column=36).value) * float(Quantity_TOTAL_WINDOWS_LICENSES_HOURS)
        Total_Precio_Final_Licencias_Windows = Total_Precio_Final_Licencias_Windows + sheet2.cell(row=m, column=37).value
      elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value <= 4): #Red Hat
        sheet2.cell(row=m, column=40).value = float(sheet2.cell(row=m, column=39).value) * float(Quantity_TOTAL_RHEL_SMALL_LICENSES)
        Total_Precio_Final_Licencias_RHS = Total_Precio_Final_Licencias_RHS + sheet2.cell(row=m, column=40).value
      elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value > 4): #windows
        sheet2.cell(row=m, column=43).value = float(sheet2.cell(row=m, column=42).value) * float(Quantity_TOTAL_RHEL_LARGE_LICENSES)
        Total_Precio_Final_Licencias_RHL = Total_Precio_Final_Licencias_RHL + sheet2.cell(row=m, column=43).value
  sheet2.cell(row=number_of_rows, column=37).value = Total_Precio_Final_Licencias_Windows
  sheet2.cell(row=number_of_rows, column=37).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=37).fill = PatternFill(start_color='bfd180', end_color='bfd180',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=40).value = Total_Precio_Final_Licencias_RHS
  sheet2.cell(row=number_of_rows, column=40).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=40).fill = PatternFill(start_color='9cd180', end_color='9cd180',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=43).value = Total_Precio_Final_Licencias_RHL
  sheet2.cell(row=number_of_rows, column=43).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=43).fill = PatternFill(start_color='d18088', end_color='d18088',fill_type = "solid")
  #Ingreso del precio final por máquina
  for m in range(2, number_of_rows):
    Total_Precio_Final_Por_Máquina = sheet2.cell(row=m, column=16).value + sheet2.cell(row=m, column=20).value + sheet2.cell(row=m, column=22).value
    if(sheet2.cell(row=m, column=12).value == '0.25 IOPS/GB'): #0.25
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=25).value
    elif (sheet2.cell(row=m, column=12).value == '4 IOPS/GB'): #4
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=28).value
    elif (sheet2.cell(row=m, column=12).value == '2 IOPS/GB'): #2
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=31).value
    elif (sheet2.cell(row=m, column=12).value == '10 IOPS/GB'): #10
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=34).value
    if('Microsoft Windows ' in sheet2.cell(row=m, column=4).value): #windows
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=37).value
    elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value <= 4): #Red Hat
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=40).value
    elif('Red Hat' in sheet2.cell(row=m, column=4).value and sheet2.cell(row=m, column=9).value > 4): #windows
      Total_Precio_Final_Por_Máquina = Total_Precio_Final_Por_Máquina + sheet2.cell(row=m, column=43).value
    sheet2.cell(row=m, column=44).value = Total_Precio_Final_Por_Máquina
    sheet2.cell(row=m, column=44).number_format = '#,##0.00$'
    Suma_Precio_Final = Suma_Precio_Final + sheet2.cell(row=m, column=44).value
    Total_Precio_Final_Por_Máquina = 0
    
  sheet2.cell(row=number_of_rows, column=44).value = Suma_Precio_Final
  sheet2.cell(row=number_of_rows, column=44).font = Font(color='ffffff', bold=True)
  sheet2.cell(row=number_of_rows, column=44).fill = PatternFill(start_color='7bb585', end_color='7bb585',fill_type = "solid")
  sheet2.cell(row=number_of_rows, column=44).number_format = '#,##0.00$'
  #N/A
  for m in range(2, number_of_rows):
    for i in range(23, 44):
      if(sheet2.cell(row=m, column=i).value==None):
        sheet2.cell(row=m, column=i).value = 'N/A'
  #Suma total de los datos constantes
  SUM_Constantes = 0
  for m in range((New_Row+1), (New_Row+11)):
      SUM_Constantes = float(SUM_Constantes) + float(sheet2.cell(m,4).value)
      sheet2.cell(m,4).number_format = '#,##0.00$'
  sheet2.cell(row=New_Row+11, column=4).value = SUM_Constantes
  sheet2.cell(row=New_Row+11, column=4).font = Font(color='000000', bold=True)
  sheet2.cell(row=New_Row+11, column=3).value = 'TOTAL'
  sheet2.cell(row=New_Row+11, column=3).font = Font(color='FF0000', bold=True)
  #Borrar datos que causan ruido
  sheet2.cell(row=120, column=5).value = None
  sheet2.cell(row=124, column=5).value = None
  #HE HECHO ESTE CAMBIO Y EL DEL FORMATO DE LOS DATOS FIJOS

  #Crear el archivo output
  Actual_Month = MonthArray[Today.month-2]
  Actual_Month = 'Reporte' + Actual_Month + '.xlsx'
  # PEQUEÑOS CAMBIOS
  #global path
  wb.save(Actual_Month)
  global nombre
  nombre = Actual_Month
  #global stream
  #stream = save_virtual_workbook(Actual_Month)
  #-------------SELECCION Y CONSTRUCCIÓN DE SOLO LOS DATOS RELEVANTES------------------------

app = Flask(__name__)

@app.route('/')
#@app.route('/home')
def home_page():
    data = {
        #HACER QUE ESTO COJA EL VALOR DE LAS VARIABLES PERO LA PREGUNTA ES DONDE LO COLOCO PARA QUE NOA FECTE LA LOGICA
        'Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS' : Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS,
        'Quantity_TOTAL_RHEL_LARGE_LICENSES' : Quantity_TOTAL_RHEL_LARGE_LICENSES,
        'Quantity_TOTAL_RHEL_SMALL_LICENSES' : Quantity_TOTAL_RHEL_SMALL_LICENSES,
        'Quantity_TOTAL_EGRESS_GB' : Quantity_TOTAL_EGRESS_GB,
        'Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS' : Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS,
        'Quantity_MAX_VCPU' : Quantity_MAX_VCPU,
        'Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS' : Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS,
        'Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS' : Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS,
        'Quantity_TOTAL_WINDOWS_LICENSES_HOURS' : Quantity_TOTAL_WINDOWS_LICENSES_HOURS,
        'Quantity_MAX_RAM_GB' : Quantity_MAX_RAM_GB,
        'Precio_Storage_025' : Precio_Storage_025,
        'Precio_Storage_4' : Precio_Storage_4,
        'Precio_Storage_2' : Precio_Storage_2,
        'Precio_Storage_10' : Precio_Storage_10
    }
    return render_template('index2.html', data = data)
@app.route('/submit', methods=['GET', 'POST'])
def submit():
    if request.method == 'POST':
        global f 
        f = request.files['file']
        f.save(secure_filename(f.filename))
        global Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS 
        Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS = request.form.get("Quantity_TOTAL_STORAGE_POINT_TWO_FIVE_IOPS_GB_HOURS")
        global Quantity_TOTAL_RHEL_LARGE_LICENSES
        Quantity_TOTAL_RHEL_LARGE_LICENSES = request.form.get("Quantity_TOTAL_RHEL_LARGE_LICENSES")
        global Quantity_TOTAL_RHEL_SMALL_LICENSES
        Quantity_TOTAL_RHEL_SMALL_LICENSES = request.form.get("Quantity_TOTAL_RHEL_SMALL_LICENSES")
        global Quantity_TOTAL_EGRESS_GB
        Quantity_TOTAL_EGRESS_GB = request.form.get("Quantity_TOTAL_EGRESS_GB")
        global Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS
        Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS = request.form.get("Quantity_TOTAL_STORAGE_TWO_IOPS_GB_HOURS")
        global Quantity_MAX_VCPU
        Quantity_MAX_VCPU = request.form.get("Quantity_MAX_VCPU")
        global Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS
        Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS = request.form.get("Quantity_TOTAL_STORAGE_TEN_IOPS_GB_HOURS")
        global Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS
        Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS = request.form.get("Quantity_TOTAL_STORAGE_FOUR_IOPS_GB_HOURS")
        global Quantity_TOTAL_WINDOWS_LICENSES_HOURS
        Quantity_TOTAL_WINDOWS_LICENSES_HOURS = request.form.get("Quantity_TOTAL_WINDOWS_LICENSES_HOURS")
        global Quantity_MAX_RAM_GB
        Quantity_MAX_RAM_GB = request.form.get("Quantity_MAX_RAM_GB")
        wholeFunction()
        return send_file(nombre, as_attachment=True)
        #return render_template("index3.html")
    elif request.method == 'GET':
        return redirect('/')
    else:
        return 'Not a valid request method for this route'

if __name__ == "__main__":
    app.run(debug=False)
#-----------------------------------------------------------